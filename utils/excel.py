import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def format_excel_worksheet(worksheet):
    """
    Excelワークシートのフォーマットを設定
    """
    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    # 列幅の自動調整
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def create_excel_receipt(data, output_path):
    """
    領収書データをExcelファイルに出力

    Parameters:
    data: 領収書情報の辞書 {"発行日": "2024/02/21", "支払先名": "株式会社〇〇", ...}
    output_path: 出力先のExcelファイルパス
    """
    try:
        # 既存のExcelファイルがあれば読み込み、なければ新規作成
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            # 既存のデータの最後のIDを取得
            last_row = ws.max_row
            start_id = ws.cell(row=last_row, column=1).value + 1 if last_row > 1 else 1
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            headers = ['ID', '発行日', '支払先名', '金額', 'インボイス番号']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
            start_id = 1

        # スタイルの設定
        alignment = Alignment(horizontal='left')

        # 新しいデータの追加
        current_row = ws.max_row + 1 if os.path.exists(output_path) else 2

        # データを追加
        ws.cell(row=current_row, column=1, value=start_id).alignment = alignment
        ws.cell(row=current_row, column=2, value=data.get('発行日', '')).alignment = alignment
        ws.cell(row=current_row, column=3, value=data.get('支払先名', '')).alignment = alignment
        ws.cell(row=current_row, column=4, value=data.get('金額', '')).alignment = alignment
        ws.cell(row=current_row, column=5, value=data.get('インボイス番号', '')).alignment = alignment

        start_id += 1

        # 列幅の自動調整
        for col in range(1, 6):
            max_length = 0
            column = get_column_letter(col)

            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # ファイルの保存
        wb.save(output_path)
        print(f"Excelファイルを保存しました: {output_path}")
        return True

    except Exception as e:
        print(f"Excelファイルの作成中にエラーが発生しました: {str(e)}")
        return False
